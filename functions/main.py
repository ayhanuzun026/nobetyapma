from firebase_functions import https_fn
from firebase_admin import initialize_app, storage
from datetime import datetime, date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
import json
from dataclasses import dataclass, field
from typing import List, Dict, Set, Optional

# OR-Tools sınıflarını dosya başında import et
from ortools_solver import (
    SolverPersonel, SolverGorev, SolverKural, SolverAtama,
    NobetSolver, HedefHesaplayici, kapasite_hesapla,
    HedefSonuc, SolverSonuc, GUN_TIPLERI, SAAT_DEGERLERI,
    normalize_id, ids_match, find_matching_id
)

# v5.0 - Frontend mantigi ile uyumlu OR-Tools
initialize_app()


# ============================================
# VERİ YAPILARI
# ============================================

@dataclass
class GorevTanim:
    id: int
    ad: str
    slot_index: int
    base_name: str = ""


@dataclass
class Personel:
    """Personel veri yapısı - Sadece gün tipi bazlı (WE/WD kaldırıldı)"""
    id: int
    ad: str
    hedef_toplam: int
    hedef_hici: int
    hedef_prs: int
    hedef_cum: int
    hedef_cmt: int
    hedef_pzr: int
    hedef_roller: Dict[str, int]

    kalan_toplam: int = 0
    kalan_hici: int = 0
    kalan_prs: int = 0
    kalan_cum: int = 0
    kalan_cmt: int = 0
    kalan_pzr: int = 0
    kalan_roller: Dict[str, int] = field(default_factory=dict)

    mazeret_gunleri: Set[int] = field(default_factory=set)
    atanan_gunler: Set[int] = field(default_factory=set)
    son_nobet_gunu: int = -999
    mazeret_sayisi: int = 0
    yillik_toplam: int = 0

    devir: Dict[str, int] = field(default_factory=dict)

    def __post_init__(self):
        self.kalan_toplam = self.hedef_toplam
        self.kalan_hici = self.hedef_hici
        self.kalan_prs = self.hedef_prs
        self.kalan_cum = self.hedef_cum
        self.kalan_cmt = self.hedef_cmt
        self.kalan_pzr = self.hedef_pzr
        self.kalan_roller = self.hedef_roller.copy()
        if isinstance(self.mazeret_gunleri, list):
            self.mazeret_gunleri = set(self.mazeret_gunleri)
        self.mazeret_sayisi = len(self.mazeret_gunleri)

    def __hash__(self):
        return hash(self.id)

    def __eq__(self, other):
        if not isinstance(other, Personel): return False
        return self.id == other.id

    def kota_kontrol(self, gun_tipi: str, gorev_adi: str, gorev_base_name: str = "") -> bool:
        """Gün tipi bazlı kota kontrolü"""
        if self.kalan_toplam <= 0: return False

        if gun_tipi == "hici" and self.kalan_hici <= 0: return False
        if gun_tipi == "persembe" and self.kalan_prs <= 0: return False
        if gun_tipi == "cuma" and self.kalan_cum <= 0: return False
        if gun_tipi == "cumartesi" and self.kalan_cmt <= 0: return False
        if gun_tipi == "pazar" and self.kalan_pzr <= 0: return False

        kontrol_adi = gorev_base_name if gorev_base_name else gorev_adi
        is_generic = (kontrol_adi.startswith("Nöbetçi") or
                      kontrol_adi.startswith("Nöbet Yeri") or
                      kontrol_adi == "Genel")

        if not is_generic:
            if self.kalan_roller.get(kontrol_adi, 0) <= 0:
                return False
        return True

    def nobet_yaz(self, gun: int, gun_tipi: str, gorev_adi: str, gorev_base_name: str = ""):
        """Nöbet yaz - sadece gün tipi kotalarını güncelle"""
        self.atanan_gunler.add(gun)
        self.son_nobet_gunu = gun
        self.kalan_toplam -= 1

        if gun_tipi == "hici":
            self.kalan_hici -= 1
        elif gun_tipi == "persembe":
            self.kalan_prs -= 1
        elif gun_tipi == "cuma":
            self.kalan_cum -= 1
        elif gun_tipi == "cumartesi":
            self.kalan_cmt -= 1
        elif gun_tipi == "pazar":
            self.kalan_pzr -= 1

        kontrol_adi = gorev_base_name if gorev_base_name else gorev_adi
        if kontrol_adi in self.kalan_roller:
            self.kalan_roller[kontrol_adi] -= 1


# ============================================
# ANA YÖNETİCİ SINIF (Sadeleştirilmiş - WE/WD kaldırıldı)
# ============================================

class NobetYoneticisi:
    """Nöbet dağıtım yöneticisi - Gün tipi bazlı sistem"""

    def __init__(self, personeller: List[Personel], gunluk_sayi: int, takvim: Dict,
                 ara_gun: int, days_in_month: int, gorev_tanimlari: List[GorevTanim],
                 kurallar: List[Dict], gorev_kisitlamalari: List[Dict] = None):
        self.personeller = personeller
        self.gunluk_sayi = gunluk_sayi
        self.takvim = takvim
        self.ara_gun = ara_gun
        self.days_in_month = days_in_month
        self.gorevler = sorted(gorev_tanimlari, key=lambda x: x.slot_index)
        self.kurallar = kurallar
        self.gorev_kisitlamalari = gorev_kisitlamalari or []
        self.cizelge = {d: [None] * len(self.gorevler) for d in range(1, days_in_month + 1)}
        self.manuel_atamalar_set = set()  # (gun, slot_idx) ciftleri - backtracking koruması

        # Mazeret istatistikleri
        self.gun_mazeret_sayisi = {}
        self._hesapla_mazeret_istatistikleri()

    def _hesapla_mazeret_istatistikleri(self):
        """Her gün için kaç kişinin mazeretli olduğunu hesapla"""
        for gun in range(1, self.days_in_month + 1):
            self.gun_mazeret_sayisi[gun] = sum(
                1 for p in self.personeller if gun in p.mazeret_gunleri
            )

    def _get_gun_tipi(self, gun: int) -> str:
        gun_adi = self.takvim.get(gun, "")
        if gun_adi == "Pazar": return "pazar"
        if gun_adi == "Cumartesi": return "cumartesi"
        if gun_adi == "Cuma": return "cuma"
        if gun_adi == "Persembe": return "persembe"
        return "hici"

    def gunleri_sirala(self) -> List[int]:
        """
        VBA MANTIĞI: En kısıtlı (mazeretli) günden başla
        1. Mazeret sayısı çok olan gün önce
        2. Zorluk sırası: Cmt (24s) > Pzr = Cum (16s) > Prş = H.İçi (8s)
        3. Tarih sırası
        """
        gun_skorlari = []
        for gun in range(1, self.days_in_month + 1):
            gun_tipi = self._get_gun_tipi(gun)
            mazeret_sayisi = self.gun_mazeret_sayisi.get(gun, 0)

            # Skor: Mazeret çok = yüksek skor (önce işlensin)
            skor = mazeret_sayisi * 1000

            # Zorluk bonusu: Cmt > Pzr = Cum > Prş = H.İçi
            if gun_tipi == "cumartesi":
                skor += 500
            elif gun_tipi == "pazar":
                skor += 400
            elif gun_tipi == "cuma":
                skor += 400
            elif gun_tipi == "persembe":
                skor += 200
            # hici = 200 (varsayılan)
            else:
                skor += 200

            gun_skorlari.append((gun, skor))

        # Yüksek skordan düşüğe sırala
        gun_skorlari.sort(key=lambda x: (-x[1], x[0]))
        return [x[0] for x in gun_skorlari]

    def gruplari_sirala(self, birlikte_kurallar: List[Dict]) -> List[Dict]:
        """
        VBA MANTIĞI: En çok mazeretli grup önce
        Grubun toplam mazeret sayısı = üyelerin mazeret sayılarının toplamı
        """

        def grup_mazeret_skoru(kural):
            toplam = 0
            for key in ['p1', 'p2', 'p3']:
                isim = kural.get(key)
                if isim:
                    p = next((x for x in self.personeller if x.ad == isim), None)
                    if p:
                        toplam += p.mazeret_sayisi
            return toplam

        return sorted(birlikte_kurallar, key=grup_mazeret_skoru, reverse=True)

    def kisi_uygun_mu(self, p: Personel, gun: int, gun_tipi: str,
                      gorev: GorevTanim, min_ara_gun: int,
                      bugun_atananlar: List[str] = None) -> bool:
        """Kişinin belirli bir güne atanıp atanamayacağını kontrol et"""
        bugun_atananlar = bugun_atananlar or []

        # Mazeret kontrolü
        if gun in p.mazeret_gunleri: return False

        # Zaten bugün atanmış mı?
        if gun in p.atanan_gunler: return False
        if p.ad in bugun_atananlar: return False

        # Kota kontrolü
        if not p.kota_kontrol(gun_tipi, gorev.ad, gorev.base_name): return False

        # Ara gün kontrolü
        for atanan_gun in p.atanan_gunler:
            if abs(gun - atanan_gun) <= min_ara_gun:
                return False

        # Görev kısıtlaması kontrolü
        for kisit in self.gorev_kisitlamalari:
            if ids_match(kisit.get('personelId'), p.id):
                kisit_gorev = kisit.get('gorevAdi')
                if kisit_gorev != gorev.ad and kisit_gorev != gorev.base_name:
                    return False

        # Ayrı tutma kuralı
        for kural in self.kurallar:
            if kural.get('tur') == 'ayri':
                p1, p2 = kural.get('p1'), kural.get('p2')
                if p.ad == p1 and p2 in bugun_atananlar: return False
                if p.ad == p2 and p1 in bugun_atananlar: return False

        return True

    def kisi_puanla(self, p: Personel, gun: int, gun_tipi: str, gorev: GorevTanim) -> float:
        """
        VBA MANTIĞI: Sıkışıklık ve panik faktörü ile puanlama
        """
        puan = 0.0

        # 1. ÖZEL GÖREV KOTASI (en yüksek öncelik)
        kontrol_adi = gorev.base_name if gorev.base_name else gorev.ad
        if kontrol_adi in p.kalan_roller and p.kalan_roller[kontrol_adi] > 0:
            puan += 5000

        # 2. SIKIŞIKLIK PUANI (VBA'dan)
        # Mazereti çok olan kişi = yüksek puan (ilk bulduğu deliğe girsin)
        puan += p.mazeret_sayisi * 100

        # 3. PANİK FAKTÖRÜ (VBA'dan)
        # Kalan hedef / Kalan boş gün oranı
        kalan_hedef = p.kalan_toplam
        kalan_bos_gun = self.days_in_month - p.mazeret_sayisi - len(p.atanan_gunler)
        if kalan_bos_gun < 1: kalan_bos_gun = 1

        panic = (kalan_hedef * 1000) / kalan_bos_gun
        puan += panic

        # 4. DEVİR ÖNCELİĞİ (önceki aydan eksik kalan)
        devir_map = {"hici": "hici", "persembe": "prs", "cuma": "cum",
                     "cumartesi": "cmt", "pazar": "pzr"}
        devir_key = devir_map.get(gun_tipi, gun_tipi)
        if p.devir.get(devir_key, 0) > 0:
            puan += 3000

        # 5. YILLIK DENGELEME (az tutan önce)
        puan -= p.yillik_toplam * 10

        # 6. AZ NÖBET TUTAN ÖNCE
        puan -= len(p.atanan_gunler) * 200

        # 7. SON NÖBETTEN UZAKLIK (homojen dağılım)
        if p.son_nobet_gunu > 0:
            puan += (gun - p.son_nobet_gunu) * 10
        else:
            puan += 500  # Hiç tutmamış = yüksek öncelik

        return puan

    def en_uygun_adayi_sec(self, gun: int, gun_tipi: str, gorev: GorevTanim,
                           min_ara_gun: int, bugun_atananlar: List[str] = None) -> Optional[Personel]:
        """En uygun adayı seç (VBA puanlama sistemiyle)"""
        bugun_atananlar = bugun_atananlar or []

        adaylar = []
        for p in self.personeller:
            if self.kisi_uygun_mu(p, gun, gun_tipi, gorev, min_ara_gun, bugun_atananlar):
                puan = self.kisi_puanla(p, gun, gun_tipi, gorev)
                adaylar.append((p, puan))

        if not adaylar:
            return None

        # En yüksek puanlı adayı seç
        adaylar.sort(key=lambda x: -x[1])
        return adaylar[0][0]

    def grup_ortak_musait_gunler(self, grup_uyeleri: List[Personel], min_ara_gun: int) -> List[int]:
        """Grubun tüm üyelerinin müsait olduğu günleri bul"""
        ortak_gunler = []
        sirali_gunler = self.gunleri_sirala()

        for gun in sirali_gunler:
            hepsi_musait = True
            for p in grup_uyeleri:
                if gun in p.mazeret_gunleri:
                    hepsi_musait = False
                    break
                if gun in p.atanan_gunler:
                    hepsi_musait = False
                    break
                # Ara gün kontrolü
                for atanan_gun in p.atanan_gunler:
                    if abs(gun - atanan_gun) <= min_ara_gun:
                        hepsi_musait = False
                        break
                if not hepsi_musait:
                    break

            if hepsi_musait:
                ortak_gunler.append(gun)

        return ortak_gunler

    def grup_dagitimi(self, min_ara_gun: int):
        """VBA MANTIĞI: Birlikte tutulacakları önce yerleştir"""
        birlikte_kurallar = [k for k in self.kurallar if k.get('tur') == 'birlikte']
        if not birlikte_kurallar:
            return

        # Grupları zorluğa göre sırala (en mazeretli önce)
        sirali_kurallar = self.gruplari_sirala(birlikte_kurallar)

        for kural in sirali_kurallar:
            grup_uyeleri = []
            for key in ['p1', 'p2', 'p3']:
                isim = kural.get(key)
                if isim:
                    p = next((x for x in self.personeller if x.ad == isim), None)
                    if p:
                        grup_uyeleri.append(p)

            if len(grup_uyeleri) < 2:
                continue

            # Grubun hedef sayısı (en az hedefi olan kadar)
            hedef_sayisi = min(p.hedef_toplam for p in grup_uyeleri)
            yazilan_sayisi = 0

            # Ortak müsait günleri bul
            ortak_gunler = self.grup_ortak_musait_gunler(grup_uyeleri, min_ara_gun)

            for gun in ortak_gunler:
                if yazilan_sayisi >= hedef_sayisi:
                    break

                gun_tipi = self._get_gun_tipi(gun)

                # Boş slotları bul
                bos_slotlar = []
                for s_idx, g_obj in enumerate(self.gorevler):
                    if self.cizelge[gun][s_idx] is None:
                        bos_slotlar.append((s_idx, g_obj))

                if len(bos_slotlar) < len(grup_uyeleri):
                    continue

                # Greedy matching ile eşleşme (O(n*m) - permutasyon yerine)
                # Önce kısıtlı kişileri (az seçenek olanları) yerleştir
                en_iyi_atama = None
                puanli_uyeler = []
                for p in grup_uyeleri:
                    uygun_slot_sayisi = sum(
                        1 for s_idx, g_obj in bos_slotlar
                        if p.kota_kontrol(gun_tipi, g_obj.ad, g_obj.base_name)
                    )
                    puanli_uyeler.append((p, uygun_slot_sayisi))

                # En az seçeneği olan kişi önce (kısıtlı olan önce yerleşsin)
                puanli_uyeler.sort(key=lambda x: x[1])

                gecici_atama = {}
                kullanilan = set()
                basarili = True

                for p, _ in puanli_uyeler:
                    yer_buldu = False
                    for s_idx, g_obj in bos_slotlar:
                        if s_idx in kullanilan:
                            continue
                        if p.kota_kontrol(gun_tipi, g_obj.ad, g_obj.base_name):
                            gecici_atama[p.id] = (s_idx, g_obj)
                            kullanilan.add(s_idx)
                            yer_buldu = True
                            break
                    if not yer_buldu:
                        basarili = False
                        break

                if basarili:
                    en_iyi_atama = gecici_atama

                if en_iyi_atama:
                    for pid, (s_idx, g_obj) in en_iyi_atama.items():
                        p = next((x for x in self.personeller if x.id == pid), None)
                        if p:
                            self.cizelge[gun][s_idx] = p.ad
                            p.nobet_yaz(gun, gun_tipi, g_obj.ad, g_obj.base_name)
                    yazilan_sayisi += 1

    def tekli_dagitim(self, min_ara_gun: int):
        """Tekli atamaları yap - backtracking destekli"""
        sirali_gunler = self.gunleri_sirala()

        for gun in sirali_gunler:
            gun_tipi = self._get_gun_tipi(gun)
            bugun_atananlar = [x for x in self.cizelge[gun] if x is not None]

            for slot_idx, gorev in enumerate(self.gorevler):
                if self.cizelge[gun][slot_idx] is not None:
                    continue

                aday = self.en_uygun_adayi_sec(gun, gun_tipi, gorev, min_ara_gun, bugun_atananlar)
                if aday:
                    self.cizelge[gun][slot_idx] = aday.ad
                    aday.nobet_yaz(gun, gun_tipi, gorev.ad, gorev.base_name)
                    bugun_atananlar.append(aday.ad)
                elif min_ara_gun > 1:
                    # Backtracking: aday bulunamazsa komşu günlerdeki atamaları geri al ve tekrar dene
                    geri_alinan = self._backtrack_komsular(gun, slot_idx, gorev, min_ara_gun)
                    if geri_alinan:
                        bugun_atananlar = [x for x in self.cizelge[gun] if x is not None]
                        aday = self.en_uygun_adayi_sec(gun, gun_tipi, gorev, min_ara_gun, bugun_atananlar)
                        if aday:
                            self.cizelge[gun][slot_idx] = aday.ad
                            aday.nobet_yaz(gun, gun_tipi, gorev.ad, gorev.base_name)
                            bugun_atananlar.append(aday.ad)

    def _backtrack_komsular(self, gun: int, slot_idx: int, gorev: GorevTanim, min_ara_gun: int) -> bool:
        """Komşu günlerdeki son 2 atamayı geri alarak yeni aday alanı aç"""
        geri_alinan = False
        # Önceki 2 gün içindeki atamalara bak
        for geri_gun in range(max(1, gun - min_ara_gun), gun):
            for geri_slot in range(len(self.gorevler) - 1, -1, -1):
                # Manuel atamayı geri alma - sabitlenmiş atamayı koru
                if (geri_gun, geri_slot) in self.manuel_atamalar_set:
                    continue
                atanan_ad = self.cizelge[geri_gun][geri_slot]
                if atanan_ad is None:
                    continue
                # Bu kişiyi bul
                kisi = None
                for p in self.personeller:
                    if p.ad == atanan_ad:
                        kisi = p
                        break
                if kisi is None:
                    continue
                # Geri al
                geri_gun_tipi = self._get_gun_tipi(geri_gun)
                geri_gorev = self.gorevler[geri_slot]
                self._atama_geri_al(kisi, geri_gun, geri_gun_tipi, geri_gorev)
                self.cizelge[geri_gun][geri_slot] = None
                geri_alinan = True
                # Sadece bir atama geri al
                return geri_alinan
        return geri_alinan

    def _atama_geri_al(self, p: Personel, gun: int, gun_tipi: str, gorev: GorevTanim):
        """Bir atamayı geri al - personelin kotalarını güncelle"""
        p.atanan_gunler.discard(gun)
        p.kalan_toplam += 1

        if gun_tipi == "hici":
            p.kalan_hici += 1
        elif gun_tipi == "persembe":
            p.kalan_prs += 1
        elif gun_tipi == "cuma":
            p.kalan_cum += 1
        elif gun_tipi == "cumartesi":
            p.kalan_cmt += 1
        elif gun_tipi == "pazar":
            p.kalan_pzr += 1

        kontrol_adi = gorev.base_name if gorev.base_name else gorev.ad
        if kontrol_adi in p.kalan_roller:
            p.kalan_roller[kontrol_adi] += 1

        # son_nobet_gunu güncelle (en son atanan güne geri dön)
        if p.atanan_gunler:
            p.son_nobet_gunu = max(p.atanan_gunler)
        else:
            p.son_nobet_gunu = -999

    def son_vurus(self):
        """Kalan boşlukları en esnek kuralla doldur (ara gün = 1)"""
        sirali_gunler = self.gunleri_sirala()

        for gun in sirali_gunler:
            gun_tipi = self._get_gun_tipi(gun)
            bugun_atananlar = [x for x in self.cizelge[gun] if x is not None]

            for slot_idx, gorev in enumerate(self.gorevler):
                if self.cizelge[gun][slot_idx] is not None:
                    continue

                aday = self.en_uygun_adayi_sec(gun, gun_tipi, gorev, 1, bugun_atananlar)
                if aday:
                    self.cizelge[gun][slot_idx] = aday.ad
                    aday.nobet_yaz(gun, gun_tipi, gorev.ad, gorev.base_name)
                    bugun_atananlar.append(aday.ad)

    def dagit(self):
        """
        Sadeleştirilmiş dağıtım - Gün tipi bazlı
        1. Grup dağıtımı (birlikte tutulacaklar)
        2. Tekli dağıtım (gün tipi kotalarına göre)
        3. Son vuruş (kalan boşluklar)
        """
        # Faz 1: Grup dağıtımı
        self.grup_dagitimi(self.ara_gun)

        # Faz 2: Tekli dağıtım
        self.tekli_dagitim(self.ara_gun)

        # Faz 3: Esnek kural ile tekrar
        if self.ara_gun > 1:
            self.tekli_dagitim(self.ara_gun - 1)

        # Faz 4: Son vuruş
        self.son_vurus()


# ============================================
# YARDIMCI FONKSİYONLAR
# ============================================

def get_days_in_month(yil, ay):
    if ay == 12:
        d = date(yil + 1, 1, 1)
    else:
        d = date(yil, ay + 1, 1)
    return (d - timedelta(days=1)).day


def gun_adi_bul(yil, ay, gun, resmi_tatiller):
    for rt in resmi_tatiller:
        if int(rt.get('gun', 0)) == gun:
            tip = rt.get('tip', '')
            if tip == "pzr": return "Pazar"
            if tip == "cmt": return "Cumartesi"
            if tip == "cum": return "Cuma"
            if tip == "prs": return "Persembe"
    dt = date(yil, ay, gun)
    gunler = ["Pazartesi", "Sali", "Carsamba", "Persembe", "Cuma", "Cumartesi", "Pazar"]
    return gunler[dt.weekday()]


def create_excel(yil, ay, yonetici: NobetYoneticisi):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nöbet Listesi"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    weekend_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    center = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Başlıklar
    headers = ["Tarih", "Gün"]
    for g in yonetici.gorevler:
        headers.append(g.ad)
    ws.append(headers)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    tr_gunler = {
        "Pazar": "Paz", "Cumartesi": "Cmt", "Cuma": "Cum",
        "Persembe": "Prş", "Pazartesi": "Pzt", "Sali": "Sal", "Carsamba": "Çar"
    }

    for gun in range(1, yonetici.days_in_month + 1):
        dt = date(yil, ay, gun)
        gun_adi_long = yonetici.takvim[gun]
        gun_kisa = tr_gunler.get(gun_adi_long, gun_adi_long)

        row_data = [dt.strftime("%d.%m.%Y"), gun_kisa]
        atamalar = yonetici.cizelge[gun]
        for kisi in atamalar:
            row_data.append(kisi if kisi else "-")
        ws.append(row_data)

        if gun_adi_long in ["Cumartesi", "Pazar"]:
            for cell in ws[ws.max_row]:
                cell.fill = weekend_fill

    # İstatistik sayfası
    ws_stat = wb.create_sheet("İstatistik")
    ws_stat.append(["Personel", "Hedef", "Gerçekleşen", "Fark", "Kalan H.İçi", "Kalan Pzr", "Mazeret Gün"])

    for p in yonetici.personeller:
        gerceklesen = len(p.atanan_gunler)
        fark = gerceklesen - p.hedef_toplam
        ws_stat.append([p.ad, p.hedef_toplam, gerceklesen, fark, p.kalan_hici, p.kalan_pzr, p.mazeret_sayisi])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ============================================
# FIREBASE FUNCTION
# ============================================

@https_fn.on_request(min_instances=0, max_instances=10, timeout_sec=540, memory=1024)
def nobet_dagit(req: https_fn.Request) -> https_fn.Response:
    headers = {
        'Access-Control-Allow-Origin': '*',
        'Access-Control-Allow-Methods': 'POST, OPTIONS',
        'Access-Control-Allow-Headers': 'Content-Type'
    }

    if req.method == 'OPTIONS':
        return https_fn.Response("", status=204, headers=headers)

    try:
        data = req.get_json(silent=True)
        if not data:
            return https_fn.Response(json.dumps({"error": "No data"}), status=400, headers=headers)

        yil = int(data.get("yil", 2025))
        ay = int(data.get("ay", 1))
        gunluk_sayi = int(data.get("gunlukSayi", 5))
        ara_gun = int(data.get("araGun", 2))
        kurallar = data.get("kurallar", [])
        gorev_kisitlamalari = data.get("gorevKisitlamalari", [])

        # Görev tanımları
        raw_gorevler = data.get("gorevTanimlari", [])
        gorev_objs = []

        if raw_gorevler and isinstance(raw_gorevler[0], dict):
            for idx, g in enumerate(raw_gorevler):
                gorev_id = g.get("id", idx)
                gorev_objs.append(GorevTanim(
                    id=gorev_id,
                    ad=g.get("ad", f"Nöbetçi {idx + 1}"),
                    slot_index=idx,
                    base_name=g.get("baseName", g.get("ad", ""))
                ))
        else:
            for idx, g_ad in enumerate(raw_gorevler):
                gorev_objs.append(GorevTanim(
                    id=idx, ad=str(g_ad), slot_index=idx, base_name=str(g_ad)
                ))

        # Eksik görevleri tamamla
        while len(gorev_objs) < gunluk_sayi:
            idx = len(gorev_objs)
            gorev_objs.append(GorevTanim(
                id=idx, ad=f"Nöbetçi {idx + 1}", slot_index=idx
            ))

        days_in_month = get_days_in_month(yil, ay)
        resmi_tatiller = data.get("resmiTatiller", [])
        takvim = {d: gun_adi_bul(yil, ay, d, resmi_tatiller) for d in range(1, days_in_month + 1)}

        # Personelleri yükle
        personeller = []
        p_list_raw = data.get("personeller", [])

        for idx, p_data in enumerate(p_list_raw):
            ad = p_data.get("ad")
            if not ad:
                continue

            hici = int(p_data.get("hici", 0))
            prs = int(p_data.get("prs", 0))
            cum = int(p_data.get("cum", 0))
            cmt = int(p_data.get("cmt", 0))
            pzr = int(p_data.get("pzr", 0))
            toplam = hici + prs + cum + cmt + pzr

            devir = p_data.get("devir", {})
            yillik_toplam = int(p_data.get("yillikToplam", 0))

            # Rol kotaları
            rol_kotalari = {}
            gorev_kotalari_raw = p_data.get("gorevKotalari", {})
            if gorev_kotalari_raw and isinstance(gorev_kotalari_raw, dict):
                for gorev_adi, kota in gorev_kotalari_raw.items():
                    try:
                        v = int(kota)
                        if v > 0:
                            rol_kotalari[gorev_adi] = v
                    except (ValueError, TypeError):
                        pass

            # Mazeretler - güvenli parse
            mazeretler = set()
            maz_raw = p_data.get("mazeretler", [])
            if isinstance(maz_raw, list):
                for x in maz_raw:
                    if x is not None:
                        try:
                            mazeretler.add(int(x))
                        except (ValueError, TypeError):
                            pass
            elif isinstance(maz_raw, dict):
                for k in maz_raw.keys():
                    try:
                        mazeretler.add(int(k))
                    except (ValueError, TypeError):
                        pass

            # Yıllık izinler ve nöbet izinleri de mazeret sayılır
            yillik_izin = p_data.get("yillikIzinler", [])
            nobet_izni = p_data.get("nobetIzinleri", [])
            if isinstance(yillik_izin, list):
                for x in yillik_izin:
                    if x is not None:
                        try:
                            mazeretler.add(int(x))
                        except (ValueError, TypeError):
                            pass
            if isinstance(nobet_izni, list):
                for x in nobet_izni:
                    if x is not None:
                        try:
                            mazeretler.add(int(x))
                        except (ValueError, TypeError):
                            pass

            personel = Personel(
                id=normalize_id(p_data.get("id", idx)),
                ad=ad,
                hedef_toplam=toplam,
                hedef_hici=hici,
                hedef_prs=prs,
                hedef_cum=cum,
                hedef_cmt=cmt,
                hedef_pzr=pzr,
                hedef_roller=rol_kotalari,
                mazeret_gunleri=mazeretler
            )
            personel.devir = devir
            personel.yillik_toplam = yillik_toplam
            personeller.append(personel)

        # Yönetici oluştur
        yonetici = NobetYoneticisi(
            personeller=personeller,
            gunluk_sayi=gunluk_sayi,
            takvim=takvim,
            ara_gun=ara_gun,
            days_in_month=days_in_month,
            gorev_tanimlari=gorev_objs,
            kurallar=kurallar,
            gorev_kisitlamalari=gorev_kisitlamalari
        )

        # Manuel atamalar
        manuel_atamalar = data.get("manuelAtamalar", [])
        for m in manuel_atamalar:
            p_ad = m.get("personel") or m.get("personelAd")
            gun = int(m.get("gun", 0))

            # BUG FIX (A): gorevId ile doğru slot'u bul (sıralama değişse bile)
            gorev_id = m.get("gorevId")
            gorev_adi = m.get("gorevAdi")
            gorev_idx = None

            if gorev_id is not None:
                # Yeni format: gorevId ile bul
                for idx, g in enumerate(gorev_objs):
                    if g.id == gorev_id:
                        gorev_idx = idx
                        break

            if gorev_idx is None and gorev_adi:
                # gorevAdi ile bul
                for idx, g in enumerate(gorev_objs):
                    if g.ad == gorev_adi:
                        gorev_idx = idx
                        break

            if gorev_idx is None:
                # Eski format fallback
                gorev_idx = int(m.get("gorevIdx", 0))

            kisi = next((p for p in personeller if p.ad == p_ad), None)
            if kisi and 1 <= gun <= days_in_month and gorev_idx is not None and gorev_idx < len(gorev_objs):
                if yonetici.cizelge[gun][gorev_idx] is None:
                    yonetici.cizelge[gun][gorev_idx] = kisi.ad
                    yonetici.manuel_atamalar_set.add((gun, gorev_idx))
                    g_obj = gorev_objs[gorev_idx]
                    gun_tipi = yonetici._get_gun_tipi(gun)
                    kisi.nobet_yaz(gun, gun_tipi, g_obj.ad, g_obj.base_name)

        # Dağıtımı yap
        yonetici.dagit()

        # Sonuçları hazırla
        sonuc_cizelge = {}
        for gun, atamalar in yonetici.cizelge.items():
            sonuc_cizelge[str(gun)] = atamalar

        kisi_ozet = []
        eksik_atamalar = []
        for p in personeller:
            gerceklesen = len(p.atanan_gunler)
            fark = p.hedef_toplam - gerceklesen
            kisi_ozet.append({
                "ad": p.ad,
                "hedef": p.hedef_toplam,
                "gerceklesen": gerceklesen,
                "fark": fark,
                "kalanHici": p.kalan_hici,
                "kalanPrs": p.kalan_prs,
                "kalanCum": p.kalan_cum,
                "kalanCmt": p.kalan_cmt,
                "kalanPzr": p.kalan_pzr
            })
            if fark > 0:
                eksik_atamalar.append({
                    "personel": p.ad,
                    "eksik": fark,
                    "detay": {
                        "hici": p.kalan_hici,
                        "prs": p.kalan_prs,
                        "cum": p.kalan_cum,
                        "cmt": p.kalan_cmt,
                        "pzr": p.kalan_pzr
                    }
                })

        # Excel oluştur ve yükle
        excel_file = create_excel(yil, ay, yonetici)
        bucket = storage.bucket()
        dosya_adi = f"sonuclar/nobet_{yil}_{ay}_{int(datetime.now().timestamp())}.xlsx"
        blob = bucket.blob(dosya_adi)
        blob.upload_from_file(
            excel_file,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        # Signed URL ile güvenli erişim (1 saat geçerli)
        signed_url = blob.generate_signed_url(
            version="v4",
            expiration=timedelta(hours=1),
            method="GET"
        )

        return https_fn.Response(
            json.dumps({
                "basari": True,
                "excelUrl": signed_url,
                "cizelge": sonuc_cizelge,
                "kisiOzet": kisi_ozet,
                "eksikAtamalar": eksik_atamalar,
                "gorevler": [g.ad for g in yonetici.gorevler]
            }),
            status=200,
            headers=headers
        )

    except Exception as e:
        import traceback
        return https_fn.Response(
            json.dumps({"error": str(e), "trace": traceback.format_exc()}),
            status=500,
            headers=headers
        )


# ============================================
# OR-TOOLS ÇÖZÜCÜ ENDPOINT'LERİ
# ============================================

# NOT: ortools_solver import'u fonksiyon içinde yapılıyor (lazy import)
# Bu sayede deploy sırasında hata oluşmuyor


def _gun_tipi_hesapla(yil: int, ay: int, gun: int, resmi_tatiller: list) -> str:
    """Gün tipini hesapla (hici, prs, cum, cmt, pzr)"""
    # Önce resmi tatillere bak
    for rt in resmi_tatiller:
        if int(rt.get('gun', 0)) == gun:
            tip = rt.get('tip', '')
            if tip == 'pzr': return 'pzr'
            if tip == 'cmt': return 'cmt'
            if tip == 'cum': return 'cum'

    # Normal gün hesapla
    dt = date(yil, ay, gun)
    weekday = dt.weekday()  # 0=Pazartesi, 6=Pazar

    if weekday == 6:  # Pazar
        return 'pzr'
    elif weekday == 5:  # Cumartesi
        return 'cmt'
    elif weekday == 4:  # Cuma
        return 'cum'
    elif weekday == 3:  # Perşembe
        return 'prs'
    else:  # Pazartesi-Çarşamba
        return 'hici'


@https_fn.on_request(min_instances=0, max_instances=10, timeout_sec=60, memory=512)
def nobet_kapasite(req: https_fn.Request) -> https_fn.Response:
    """
    Hedef önizleme için kapasite hesapla.
    OR-Tools çözmeden sadece matematiksel hesaplama yapar.
    """
    # ✅ Import kaldırıldı - dosyanın başındaki sınıflar kullanılacak

    headers = {
        'Access-Control-Allow-Origin': '*',
        'Access-Control-Allow-Methods': 'POST, OPTIONS',
        'Access-Control-Allow-Headers': 'Content-Type'
    }

    if req.method == 'OPTIONS':
        return https_fn.Response("", status=204, headers=headers)

    try:
        data = req.get_json(silent=True)
        if not data:
            return https_fn.Response(
                json.dumps({"error": "No data"}),
                status=400, headers=headers
            )

        yil = int(data.get("yil", 2025))
        ay = int(data.get("ay", 1))
        slot_sayisi = int(data.get("slotSayisi", 5))
        resmi_tatiller = data.get("resmiTatiller", [])

        gun_sayisi = get_days_in_month(yil, ay)

        # Gün tiplerini hesapla
        gun_tipleri = {}
        for g in range(1, gun_sayisi + 1):
            gun_tipleri[g] = _gun_tipi_hesapla(yil, ay, g, resmi_tatiller)

        # Personelleri parse et
        personeller = []
        for p_data in data.get("personeller", []):
            if not p_data.get("ad"):
                continue

            # Mazeretleri birleştir
            mazeretler = set()
            for key in ['mazeretler', 'yillikIzinler', 'nobetIzinleri']:
                raw = p_data.get(key, [])
                if isinstance(raw, list):
                    mazeretler.update(int(x) for x in raw if x)

            # Yıllık gerçekleşen
            yillik_gerceklesen = {}
            yg_raw = p_data.get("yillikGerceklesen", {})
            if isinstance(yg_raw, dict):
                for key, val in yg_raw.items():
                    try:
                        yillik_gerceklesen[key] = int(val)
                    except (ValueError, TypeError):
                        yillik_gerceklesen[key] = 0

            personeller.append(SolverPersonel(
                id=normalize_id(p_data.get("id", len(personeller))),
                ad=p_data.get("ad"),
                mazeret_gunleri=mazeretler,
                kisitli_gorev=p_data.get("kisitliGorev"),
                yillik_gerceklesen=yillik_gerceklesen
            ))

        # Kapasite hesapla
        sonuc = kapasite_hesapla(
            gun_sayisi=gun_sayisi,
            gun_tipleri=gun_tipleri,
            personeller=personeller,
            slot_sayisi=slot_sayisi
        )

        return https_fn.Response(
            json.dumps({"basari": True, **sonuc}),
            status=200,
            headers=headers
        )

    except Exception as e:
        import traceback
        return https_fn.Response(
            json.dumps({"error": str(e), "trace": traceback.format_exc()}),
            status=500,
            headers=headers
        )


@https_fn.on_request(min_instances=0, max_instances=5, timeout_sec=300, memory=1024)
def nobet_hedef_hesapla(req: https_fn.Request) -> https_fn.Response:
    """
    OR-Tools ile optimal hedef hesapla - v2.2
    """
    # ✅ Import kaldırıldı - dosyanın başındaki sınıflar kullanılacak

    headers = {
        'Access-Control-Allow-Origin': '*',
        'Access-Control-Allow-Methods': 'POST, OPTIONS',
        'Access-Control-Allow-Headers': 'Content-Type'
    }

    if req.method == 'OPTIONS':
        return https_fn.Response("", status=204, headers=headers)

    try:
        data = req.get_json(silent=True)
        if not data:
            return https_fn.Response(
                json.dumps({"error": "No data"}),
                status=400,
                headers=headers
            )

        # Parametreleri al
        gun_sayisi = data.get("gunSayisi", 31)
        gun_tipleri_raw = data.get("gunTipleri", {})
        gun_tipleri = {int(k): v for k, v in gun_tipleri_raw.items()}
        ara_gun = data.get("araGun", 2)
        saat_degerleri = data.get("saatDegerleri", None)

        # Personelleri dönüştür
        personeller = []
        for p_data in data.get("personeller", []):
            # ID'yi int'e standardize et
            raw_id = p_data.get("id", len(personeller))
            try:
                pid = int(float(raw_id))
            except (ValueError, TypeError):
                pid = len(personeller)

            # Mazeretleri birleştir (tüm kaynaklardan)
            mazeret_set = set()
            for key in ['mazeretler', 'yillikIzinler', 'nobetIzinleri']:
                raw = p_data.get(key, [])
                if isinstance(raw, list):
                    for m in raw:
                        if m is not None:
                            try:
                                mazeret_set.add(int(m))
                            except (ValueError, TypeError):
                                pass
                elif isinstance(raw, dict):
                    for k in raw.keys():
                        try:
                            mazeret_set.add(int(k))
                        except (ValueError, TypeError):
                            pass

            # Yıllık gerçekleşen
            yillik_gerceklesen = {}
            yg_raw = p_data.get("yillikGerceklesen", {})
            if isinstance(yg_raw, dict):
                for key, val in yg_raw.items():
                    try:
                        yillik_gerceklesen[key] = int(val)
                    except (ValueError, TypeError):
                        yillik_gerceklesen[key] = 0

            personeller.append(SolverPersonel(
                id=pid,
                ad=p_data.get("ad", ""),
                mazeret_gunleri=mazeret_set,
                kisitli_gorev=p_data.get("kisitliGorev"),
                yillik_gerceklesen=yillik_gerceklesen
            ))

        # Görevleri dönüştür
        gorevler = []
        for idx, g_data in enumerate(data.get("gorevler", [])):
            gorevler.append(SolverGorev(
                id=g_data.get("id", idx),
                ad=g_data.get("ad", f"Görev {idx}"),
                slot_idx=idx,
                base_name=g_data.get("baseName", ""),
                exclusive=g_data.get("exclusive", False)
            ))

        # Birlikte kurallarını dönüştür
        birlikte_kurallar = []
        for k_data in data.get("kurallar", []):
            if k_data.get("tur") == "birlikte":
                # Kişileri ID veya isimden bul
                kisiler = []
                for key in ['p1', 'p2', 'p3', 'kisiler']:
                    val = k_data.get(key)
                    if val is None:
                        continue
                    
                    # kisiler array olabilir
                    if key == 'kisiler' and isinstance(val, list):
                        for v in val:
                            if isinstance(v, (int, float)):
                                kisiler.append(int(float(v)))
                            elif isinstance(v, str):
                                try:
                                    kisiler.append(int(float(v)))
                                except (ValueError, TypeError):
                                    for p in personeller:
                                        if p.ad == v:
                                            kisiler.append(p.id)
                                            break
                    elif isinstance(val, (int, float)):
                        kisiler.append(int(float(val)))
                    elif isinstance(val, str):
                        try:
                            kisiler.append(int(float(val)))
                        except (ValueError, TypeError):
                            for p in personeller:
                                if p.ad == val:
                                    kisiler.append(p.id)
                                    break
                
                if len(kisiler) >= 2:
                    birlikte_kurallar.append(SolverKural(
                        tur="birlikte",
                        kisiler=kisiler
                    ))

        # Görev kısıtlamalarını dönüştür
        gorev_kisitlamalari = {}
        for k_data in data.get("gorevKisitlamalari", []):
            pid = k_data.get("personelId")
            gorev_adi = k_data.get("gorevAdi")
            if pid is not None and gorev_adi:
                try:
                    gorev_kisitlamalari[int(float(pid))] = gorev_adi
                except (ValueError, TypeError):
                    pass

        # Manuel atamaları dönüştür
        manuel_atamalar = []
        for m_data in data.get("manuelAtamalar", []):
            # Personel ID bul (isim veya ID olabilir)
            p_id = None
            p_ad = m_data.get("personel") or m_data.get("personelAd")
            p_raw_id = m_data.get("personelId")
            
            if p_raw_id is not None:
                try:
                    p_id = int(float(p_raw_id))
                except (ValueError, TypeError):
                    pass
            
            if p_id is None and p_ad:
                for p in personeller:
                    if p.ad == p_ad:
                        p_id = p.id
                        break
            
            if p_id is None:
                continue
            
            gun = m_data.get("gun")
            if gun is None:
                continue
            try:
                gun = int(gun)
            except (ValueError, TypeError):
                continue
            
            manuel_atamalar.append(SolverAtama(
                personel_id=p_id,
                gun=gun,
                slot_idx=int(m_data.get("slotIdx", 0)),
                gorev_adi=m_data.get("gorevAdi", "")
            ))

        # Hedef hesapla
        hesaplayici = HedefHesaplayici(
            gun_sayisi=gun_sayisi,
            gun_tipleri=gun_tipleri,
            personeller=personeller,
            gorevler=gorevler,
            birlikte_kurallar=birlikte_kurallar,
            gorev_kisitlamalari=gorev_kisitlamalari,
            manuel_atamalar=manuel_atamalar,
            ara_gun=ara_gun,
            saat_degerleri=saat_degerleri
        )

        sonuc = hesaplayici.hesapla()

        return https_fn.Response(
            json.dumps({
                "basari": sonuc.basarili,
                "hedefler": sonuc.hedefler,
                "birlikteAtamalar": sonuc.birlikte_atamalar,
                "gorevKotalari": sonuc.gorev_kotalari,
                "istatistikler": sonuc.istatistikler,
                "mesaj": sonuc.mesaj
            }),
            status=200,
            headers=headers
        )

    except Exception as e:
        import traceback
        return https_fn.Response(
            json.dumps({"error": str(e), "trace": traceback.format_exc()}),
            status=500,
            headers=headers
        )


@https_fn.on_request(min_instances=0, max_instances=5, timeout_sec=540, memory=2048)
def nobet_coz(req: https_fn.Request) -> https_fn.Response:
    """
    OR-Tools CP-SAT ile nöbet çöz.
    Görev kotaları + Gün tipi kotaları dahil.
    """
    # ✅ Import kaldırıldı - dosyanın başındaki sınıflar kullanılacak

    headers = {
        'Access-Control-Allow-Origin': '*',
        'Access-Control-Allow-Methods': 'POST, OPTIONS',
        'Access-Control-Allow-Headers': 'Content-Type'
    }

    if req.method == 'OPTIONS':
        return https_fn.Response("", status=204, headers=headers)

    try:
        data = req.get_json(silent=True)
        if not data:
            return https_fn.Response(json.dumps({"error": "No data"}), status=400, headers=headers)

        yil = int(data.get("yil", 2025))
        ay = int(data.get("ay", 1))
        slot_sayisi = int(data.get("slotSayisi", 6))
        ara_gun = int(data.get("araGun", 2))
        max_sure = int(data.get("maxSure", 300))
        resmi_tatiller = data.get("resmiTatiller", [])
        saat_degerleri = data.get("saatDegerleri", None)

        gun_sayisi = get_days_in_month(yil, ay)

        # Gün tiplerini hesapla
        gun_tipleri = {}
        for g in range(1, gun_sayisi + 1):
            gun_tipleri[g] = _gun_tipi_hesapla(yil, ay, g, resmi_tatiller)

        # Kişi bazlı exclusive kontrolü için görev kısıtlamalarını önce al
        gorev_kisitlamalari_raw = data.get("gorevKisitlamalari", [])
        exclusive_gorevler = set()  # Hangi görevler exclusive?
        for k in gorev_kisitlamalari_raw:
            # Kişi bazlı exclusive: k.exclusive true ise bu görev exclusive
            if k.get("exclusive", False):
                exclusive_gorevler.add(k.get("gorevAdi"))
        
        # Görevleri parse et - BASE_NAME ÖNEMLİ!
        gorevler = []
        for idx, g_data in enumerate(data.get("gorevler", [])):
            if isinstance(g_data, dict):
                gorev_ad = g_data.get("ad", f"Nöbetçi {idx + 1}")
                # base_name: "AMELİYATHANE #1" -> "AMELİYATHANE"
                base_name = g_data.get("baseName", gorev_ad.split(" #")[0] if " #" in gorev_ad else gorev_ad)
                # Kişi bazlı exclusive: gorev adı veya base_name exclusive_gorevler'da mı?
                exclusive = g_data.get("exclusive", False) or gorev_ad in exclusive_gorevler or base_name in exclusive_gorevler
            else:
                gorev_ad = str(g_data)
                base_name = gorev_ad.split(" #")[0] if " #" in gorev_ad else gorev_ad
                exclusive = gorev_ad in exclusive_gorevler or base_name in exclusive_gorevler

            gorevler.append(SolverGorev(
                id=idx,
                ad=gorev_ad,
                slot_idx=idx,
                base_name=base_name,
                exclusive=exclusive
            ))

        # Eksik görevleri tamamla
        while len(gorevler) < slot_sayisi:
            idx = len(gorevler)
            gorevler.append(SolverGorev(id=idx, ad=f"Nöbetçi {idx + 1}", slot_idx=idx, base_name=f"Nöbetçi {idx + 1}"))

        # Personelleri parse et
        personeller = []
        for p_data in data.get("personeller", []):
            if not p_data.get("ad"):
                continue

            # ID'yi güvenli şekilde normalize et
            raw_id = p_data.get("id", len(personeller))
            pid = normalize_id(raw_id)

            # Mazeretleri birleştir
            mazeretler = set()
            for key in ['mazeretler', 'yillikIzinler', 'nobetIzinleri']:
                raw = p_data.get(key, [])
                if isinstance(raw, list):
                    for x in raw:
                        if x:
                            try:
                                mazeretler.add(int(x))
                            except (ValueError, TypeError):
                                pass

            # Görev kısıtlaması - int ID ile güvenli karşılaştırma
            kisitli_gorev = None
            for k in data.get("gorevKisitlamalari", []):
                k_pid = k.get("personelId", -1)
                try:
                    if normalize_id(k_pid) == pid:
                        raw_gorev_adi = k.get("gorevAdi")
                        # Frontend slot adı gönderebilir (ör: "AMEL #1"), base_name'e çevir
                        kisitli_gorev = raw_gorev_adi
                        for g in gorevler:
                            if g.ad == raw_gorev_adi and g.base_name:
                                kisitli_gorev = g.base_name
                                break
                        break
                except (ValueError, TypeError):
                    pass

            # Gün tipi hedefleri
            hedef_tipler = {}
            for tip in ['hici', 'prs', 'cum', 'cmt', 'pzr']:
                val = p_data.get(tip)
                if val is not None:
                    try:
                        hedef_tipler[tip] = int(val)
                    except (ValueError, TypeError):
                        hedef_tipler[tip] = 0

            # Görev kotaları
            gorev_kotalari = {}
            gk_raw = p_data.get("gorevKotalari", {})
            if isinstance(gk_raw, dict):
                for gorev_adi, kota in gk_raw.items():
                    try:
                        gorev_kotalari[gorev_adi] = int(kota)
                    except (ValueError, TypeError):
                        pass

            # YENİ: Yıllık gerçekleşen (önceki ayların toplamı)
            yillik_gerceklesen = {}
            yg_raw = p_data.get("yillikGerceklesen", {})
            if isinstance(yg_raw, dict):
                for key, val in yg_raw.items():
                    try:
                        yillik_gerceklesen[key] = int(val)
                    except (ValueError, TypeError):
                        yillik_gerceklesen[key] = 0

            personeller.append(SolverPersonel(
                id=pid,
                ad=p_data.get("ad"),
                mazeret_gunleri=mazeretler,
                kisitli_gorev=kisitli_gorev,
                hedef_tipler=hedef_tipler,
                gorev_kotalari=gorev_kotalari,
                yillik_gerceklesen=yillik_gerceklesen
            ))

        # Kuralları parse et
        kurallar = []
        for k_data in data.get("kurallar", []):
            tur = k_data.get("tur")
            if tur not in ['ayri', 'birlikte']:
                continue

            kisiler = []
            # kisiler array'ini kontrol et (frontend bu formatı kullanıyor)
            kisiler_raw = k_data.get("kisiler", [])
            if isinstance(kisiler_raw, list):
                for v in kisiler_raw:
                    if isinstance(v, (int, float)):
                        kisiler.append(int(float(v)))
                    elif isinstance(v, str):
                        try:
                            kisiler.append(int(float(v)))
                        except (ValueError, TypeError):
                            for p in personeller:
                                if p.ad == v:
                                    kisiler.append(p.id)
                                    break
            
            # Eski format için de kontrol et (p1, p2, p3)
            if len(kisiler) == 0:
                for key in ['p1', 'p2', 'p3']:
                    pid = k_data.get(key)
                    if pid is not None:
                        if isinstance(pid, (int, float)):
                            kisiler.append(int(float(pid)))
                        elif isinstance(pid, str):
                            try:
                                kisiler.append(int(float(pid)))
                            except (ValueError, TypeError):
                                for p in personeller:
                                    if p.ad == pid:
                                        kisiler.append(p.id)
                                        break

            if len(kisiler) >= 2:
                kurallar.append(SolverKural(tur=tur, kisiler=kisiler))

        # Manuel atamaları parse et
        manuel_atamalar = []
        for m_data in data.get("manuelAtamalar", []):
            gun = int(m_data.get("gun", 0))
            if gun < 1 or gun > gun_sayisi:
                continue

            # Personel ID bul
            p_ad = m_data.get("personel") or m_data.get("personelAd")
            p_id = None
            for p in personeller:
                if p.ad == p_ad:
                    p_id = p.id
                    break

            if p_id is None:
                continue

            # Slot bul
            gorev_adi = m_data.get("gorevAdi")
            slot_idx = None
            for g in gorevler:
                if g.ad == gorev_adi:
                    slot_idx = g.slot_idx
                    break

            if slot_idx is None:
                slot_idx = int(m_data.get("gorevIdx", 0))

            if slot_idx < len(gorevler):
                manuel_atamalar.append(SolverAtama(
                    personel_id=p_id,
                    gun=gun,
                    slot_idx=slot_idx
                ))

        # Hedefleri hazırla - HedefHesaplayici ile tutarlı hesapla
        # Frontend'den gelen hedefler varsa kullan, yoksa HedefHesaplayici ile hesapla
        frontend_hedefleri_var = any(
            p.hedef_tipler and sum(p.hedef_tipler.values()) > 0
            for p in personeller
        )

        if frontend_hedefleri_var:
            # Frontend hedefleri kullan
            hedefler = {}
            toplam_slot = gun_sayisi * len(gorevler)
            kisi_sayisi = len(personeller)
            kisi_basi_hedef = toplam_slot // kisi_sayisi if kisi_sayisi > 0 else 0
            kalan = toplam_slot % kisi_sayisi if kisi_sayisi > 0 else 0

            for idx, p in enumerate(personeller):
                if p.hedef_tipler and sum(p.hedef_tipler.values()) > 0:
                    hedef_toplam = sum(p.hedef_tipler.values())
                else:
                    musait_gun = gun_sayisi - len(p.mazeret_gunleri)
                    hedef_toplam = min(kisi_basi_hedef + (1 if idx < kalan else 0), musait_gun)

                hedefler[p.id] = {
                    'hedef_toplam': hedef_toplam,
                    'hedef_tipler': p.hedef_tipler or {},
                    'gorev_kotalari': p.gorev_kotalari or {}
                }
        else:
            # HedefHesaplayici ile hesapla (nobet_hedef_hesapla ile aynı)
            birlikte_kurallar = [k for k in kurallar if k.tur == 'birlikte']
            gorev_kisitlamalari_dict = {}
            for k_data in data.get("gorevKisitlamalari", []):
                pid_raw = k_data.get("personelId")
                gorev_adi = k_data.get("gorevAdi")
                if pid_raw is not None and gorev_adi:
                    try:
                        gorev_kisitlamalari_dict[int(float(pid_raw))] = gorev_adi
                    except (ValueError, TypeError):
                        pass

            hesaplayici = HedefHesaplayici(
                gun_sayisi=gun_sayisi,
                gun_tipleri=gun_tipleri,
                personeller=personeller,
                gorevler=gorevler,
                birlikte_kurallar=birlikte_kurallar,
                gorev_kisitlamalari=gorev_kisitlamalari_dict,
                manuel_atamalar=manuel_atamalar,
                ara_gun=ara_gun,
                saat_degerleri=saat_degerleri
            )
            hesap_sonuc = hesaplayici.hesapla()
            hedefler = {}
            for h in hesap_sonuc.hedefler:
                pid = h.get('id')
                hedefler[pid] = {
                    'hedef_toplam': h.get('hedef_toplam', 0),
                    'hedef_tipler': h.get('hedef_tipler', {}),
                    'gorev_kotalari': h.get('gorev_kotalari', {})
                }

        # Kademeli çözüm: ara_gun azaltarak fallback (0 dahil)
        sonuc = None
        kullanilan_ara_gun = ara_gun
        for dene_ara_gun in range(ara_gun, -1, -1):
            solver = NobetSolver(
                gun_sayisi=gun_sayisi,
                gun_tipleri=gun_tipleri,
                personeller=personeller,
                gorevler=gorevler,
                kurallar=kurallar,
                manuel_atamalar=manuel_atamalar,
                hedefler=hedefler,
                ara_gun=dene_ara_gun,
                max_sure_saniye=max_sure
            )

            sonuc = solver.coz()
            kullanilan_ara_gun = dene_ara_gun
            if sonuc.basarili:
                break

        # Döngü 0'a kadar indigi icin ayrıca ara_gun=0 denemesine gerek yok
        # Ama sonuc hala None ise (personel yok vb.) güvenli dönüş yap
        if sonuc is None:
            sonuc = SolverSonuc(
                basarili=False,
                atamalar=[],
                istatistikler={'status': 'NO_SOLUTION', 'ara_gun': ara_gun},
                sure_ms=0,
                mesaj="Çözüm üretilemedi - parametre hatası olabilir"
            )
            kullanilan_ara_gun = ara_gun

        # Fallback bilgisini mesaja ekle
        if kullanilan_ara_gun != ara_gun and sonuc.basarili:
            sonuc = SolverSonuc(
                basarili=sonuc.basarili,
                atamalar=sonuc.atamalar,
                istatistikler={**sonuc.istatistikler,
                               'fallback_ara_gun': kullanilan_ara_gun,
                               'istenen_ara_gun': ara_gun},
                sure_ms=sonuc.sure_ms,
                mesaj=f"{sonuc.mesaj} (ara_gun {ara_gun}->{kullanilan_ara_gun} gevsetildi)"
            )

        # Çizelge formatına dönüştür (frontend uyumluluğu için)
        cizelge = {}
        for g in range(1, gun_sayisi + 1):
            cizelge[str(g)] = [None] * len(gorevler)

        for atama in sonuc.atamalar:
            gun = atama['gun']
            slot = atama['slot_idx']
            ad = atama['personel_ad']
            cizelge[str(gun)][slot] = ad

        # Debug: Hedef eşleşme kontrolü
        hedef_debug = []
        for p in personeller:
            h = hedefler.get(p.id, {})
            hedef_debug.append({
                'id': p.id,
                'ad': p.ad,
                'hedef_toplam': h.get('hedef_toplam', 0),
                'hedef_tipler': h.get('hedef_tipler', {}),
                'mazeret_sayisi': len(p.mazeret_gunleri)
            })

        return https_fn.Response(
            json.dumps({
                "basari": sonuc.basarili,
                "mesaj": sonuc.mesaj,
                "sureMs": sonuc.sure_ms,
                "cizelge": cizelge,
                "atamalar": sonuc.atamalar,
                "istatistikler": sonuc.istatistikler,
                "gorevler": [g.ad for g in gorevler],
                "hedefDebug": hedef_debug
            }),
            status=200,
            headers=headers
        )

    except Exception as e:
        import traceback
        return https_fn.Response(
            json.dumps({"error": str(e), "trace": traceback.format_exc()}),
            status=500,
            headers=headers
        )
