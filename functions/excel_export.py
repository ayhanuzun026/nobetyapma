"""
Excel export fonksiyonu — OR-Tools solver sonucuyla calisan versiyon.
"""

from datetime import date
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from utils import gun_adi_bul, normalize_id


def create_excel(yil, ay, cizelge, gorevler, personeller, hedefler, gun_sayisi,
                 resmi_tatiller=None):
    """OR-Tools sonucundan Excel raporu uretir.

    Args:
        yil: Yil
        ay: Ay
        cizelge: {str(gun): [personel_ad, ...]} formati
        gorevler: List[SolverGorev]
        personeller: List[SolverPersonel]
        hedefler: {personel_id: {hedef_toplam, hedef_tipler, ...}}
        gun_sayisi: Aydaki gun sayisi
        resmi_tatiller: Resmi tatil listesi
    """
    resmi_tatiller = resmi_tatiller or []

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nobet Listesi"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    weekend_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    center = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ["Tarih", "Gun"]
    for g in gorevler:
        headers.append(g.ad)
    ws.append(headers)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    tr_gunler = {
        "Pazar": "Paz", "Cumartesi": "Cmt", "Cuma": "Cum",
        "Persembe": "Prs", "Pazartesi": "Pzt", "Sali": "Sal", "Carsamba": "Car"
    }

    for gun in range(1, gun_sayisi + 1):
        dt = date(yil, ay, gun)
        gun_adi_long = gun_adi_bul(yil, ay, gun, resmi_tatiller)
        gun_kisa = tr_gunler.get(gun_adi_long, gun_adi_long)

        row_data = [dt.strftime("%d.%m.%Y"), gun_kisa]
        slotlar = cizelge.get(str(gun), [None] * len(gorevler))
        for kisi in slotlar:
            row_data.append(kisi if kisi else "-")
        ws.append(row_data)

        if gun_adi_long in ["Cumartesi", "Pazar"]:
            for cell in ws[ws.max_row]:
                cell.fill = weekend_fill

    # Istatistik sayfasi
    ws_stat = wb.create_sheet("Istatistik")
    ws_stat.append(["Personel", "Hedef", "Gerceklesen", "Fark", "Mazeret Gun"])

    kisi_sayac = {}
    for gun_str, slotlar in cizelge.items():
        for personel_ad in slotlar:
            if personel_ad:
                kisi_sayac[personel_ad] = kisi_sayac.get(personel_ad, 0) + 1

    for p in personeller:
        pid = normalize_id(p.id)
        h = hedefler.get(p.id) or hedefler.get(pid) or {}
        hedef_toplam = h.get('hedef_toplam', 0)
        gerceklesen = kisi_sayac.get(p.ad, 0)
        fark = gerceklesen - hedef_toplam
        mazeret_sayisi = len(p.mazeret_gunleri) if hasattr(p, 'mazeret_gunleri') else 0
        ws_stat.append([p.ad, hedef_toplam, gerceklesen, fark, mazeret_sayisi])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
